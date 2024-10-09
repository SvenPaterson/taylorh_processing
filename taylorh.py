import csv, os
import pandas as pd
import matplotlib.pyplot as plt

from datetime import datetime
from time import sleep

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, NamedStyle
import openpyxl

import tkinter as tk
from tkinter import filedialog, messagebox

def get_data_paths():
    # Allow user to select either a folder containing subfolders or a single CSV file
    root = tk.Tk()
    root.withdraw()
    docs_path = os.path.expanduser('~/Desktop')
    folder_path = filedialog.askdirectory(
        initialdir=docs_path,
        title="Choose a folder containing subfolders with CSV files or a single folder with a CSV file"
    )
    root.destroy()

    if not folder_path:
        return None, None

    # Check if the folder contains subfolders or a single CSV file
    csv_paths = []
    for root_dir, sub_dirs, files in os.walk(folder_path):
        if not sub_dirs:  # If there are no subdirectories, look for a single CSV file
            for file in files:
                if file.endswith('.csv'):
                    csv_paths.append(os.path.join(root_dir, file))
            if csv_paths:
                return csv_paths, folder_path
            else:
                tk.Tk().withdraw()
                messagebox.showwarning("No CSV File Found", "The selected folder does not contain any CSV files.")
                return None, None
        else:  # If subdirectories exist, continue searching for CSV files in subfolders
            for file in files:
                if file.endswith('.csv'):
                    csv_paths.append(os.path.join(root_dir, file))
    
    if not csv_paths:
        tk.Tk().withdraw()
        messagebox.showwarning("No CSV Files Found", "No CSV files were found in the selected folder or its subfolders.")
        return None, None

    return csv_paths, folder_path

def get_params(file_path):
    """ Reads the 2nd line of the CSV output file for the list of parameters"""
    with open(file_path, 'r', encoding='ISO-8859-1') as csvfile:
        reader = csv.reader(csvfile)
        next(reader)
        params = next(reader)
        return params


def get_units(file_path):
    """ Reads the 3rd line of the CSV output file for the measurement units """
    with open(file_path, 'r', encoding='ISO-8859-1') as csvfile:
        reader = csv.reader(csvfile)
        next(reader)
        next(reader)
        units = next(reader)
    return units


def measurement_output_to_df(file_path):
    if file_path is None:
        print("File selection was canceled. Closing in 5 seconds.")
        sleep(10)
        return None
    try:
        params = get_params(file_path)
        header = ["Date", "Time", "Measurement File"] + params[3:]
        df = pd.read_csv(file_path, skiprows=3, names=header,
                         parse_dates=[["Date", "Time"]],
                         encoding='ISO-8859-1')
        units = get_units(file_path)
        for i, column_name in enumerate(df.columns[2:]):
            df.rename(columns={column_name: column_name + ", " + units[3 + i]},
                      inplace=True)
        df.drop(columns=df.columns[-1], inplace=True)
        return df
    except Exception as e:
        tk.Tk().withdraw()
        messagebox.showerror("Error", f"Failed to process the file: {e}" +
                             "\nPlease ensure the chosen file was produced " +
                             "by the Mountains software 'apply a template' " +
                             "function")
        return None

def convert_to_metric(df):
    for column_name in df.columns[2:]:
        if "µin" in column_name and "Rmr" not in column_name:
            df[column_name] = df[column_name] * 0.0254
            new_column_name = column_name.replace("µin", "µm")
            df.rename(columns={column_name: new_column_name}, inplace=True)

def convert_to_standard(df):
    df_standard = df.copy()
    for column_name in df_standard.columns:
        if "µm" in column_name and "Rmr" not in column_name:
            df_standard[column_name] = df_standard[column_name] / 0.0254
            new_column_name = column_name.replace("µm", "µin")
            df_standard.rename(columns={column_name: new_column_name},
                               inplace=True)
    return df_standard

def generate_stats_table(df, sig_figs=3):
    measurement_columns = df.columns[3:]
    stats = df[measurement_columns].agg(['mean', 'std'])

    # Format numbers to the desired number of significant figures
    format_str = '{:.' + str(sig_figs) + 'g}'
    stats = stats.map(lambda x: format_str.format(x))

    for col in stats.columns:
        stats[col] = pd.to_numeric(stats[col])
    return stats

def save_to_excel(df, filename, processed_files, constant_width):
    # gap size & number styles
    decimal_style = NamedStyle(name="three_decimal", number_format="0.000")
    gap = 3

    df_standard = convert_to_standard(df)

    # Save the DataFrame to an Excel file using pandas
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        # Get the max row of the written DataFrame to determine where to start appending
        workbook = writer.book
        sheet = writer.sheets['Sheet1']
        metric_table_size = sheet.max_row

    # Start appending the standard DataFrame using openpyxl for more control
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook['Sheet1']
    # Calculate the row to start appending the standard DataFrame
    start_row = metric_table_size + gap + 1

    # Append the standard DataFrame to the Excel sheet
    for r_idx, (index, row) in enumerate(df_standard.iterrows(), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            sheet.cell(row=r_idx, column=c_idx, value=value)

    # Insert a new column
    sheet.insert_cols(1)

    # Set values for the first two cells in the metric table
    sheet['A1'] = ''  # Blank cell
    sheet['A2'] = 'AVG'
    sheet['A3'] = 'STDEV'
    cell = sheet.cell(column=1, row=start_row - 1)
    cell.value = "FREEDOM UNITS!"
    cell.font = Font(bold=True)
    sheet.cell(column=1, row=start_row).value = 'AVG'
    sheet.cell(column=1, row=start_row + 1).value = 'STDEV'
    standard_headers = df_standard.columns.tolist()
    for col in range(len(standard_headers)):
        cell = sheet.cell(column=col + 2, row=6)
        cell.value = standard_headers[col]
        cell.font = Font(bold=True)

    # Add the list of processed file paths
    processed_files_start_row = sheet.max_row + gap + 1
    sheet.cell(row=processed_files_start_row, column=2, value="Processed Files:")
    sheet.cell(row=processed_files_start_row, column=2).font = Font(bold=True)

    # Write the file paths
    for i, file in enumerate(processed_files, start=processed_files_start_row + 1):
        sheet.cell(row=i, column=1, value=file)

    # Set column width and cell formatting
    for col in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = constant_width
        for row in range(1, processed_files_start_row):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = Alignment(wrapText=True)
            if row == 2 or row == 7 or row == 3 or row == 8:
                cell.style = decimal_style

    # Save the workbook
    workbook.save(filename)

def main():
    file_paths, parent_folder = get_data_paths()
    if file_paths is None:
        print("No files selected. Exiting...")
        return

    for file_path in file_paths:
        print(f"Processing file: {file_path}")
        output = measurement_output_to_df(file_path)
        if output is None:
            continue
        convert_to_metric(output)
        stats_table = generate_stats_table(output)
        constant_width = 20
        now = datetime.now()
        dt_suffix = now.strftime("%Y%m%d_%H%M%S")

        # Save the result in the same subfolder as the original CSV file
        save_folder = os.path.dirname(file_path) if parent_folder else '.'
        save_path = os.path.join(save_folder, f'surf_stats_data_{dt_suffix}.xlsx')
        processed_files = output['Measurement File'].tolist()
        save_to_excel(stats_table, save_path, processed_files, constant_width)

    print("Processing completed for all files. Closing in 10 seconds...")
    sleep(10)

if __name__ == "__main__":
    main()